import signal, sys
import mysql.connector
import openpyxl
from mysql.connector import IntegrityError

def cortarPrograma(sig, frame):
    print("\n\n[!]Saliendo ...\n")
    sys.exit(1)


# Ctlr + C
signal.signal(signal.SIGINT, cortarPrograma)
conexion = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="permisos",
    port=3307,
)
cursor = conexion.cursor()

wb = openpyxl.load_workbook("docs/excel-cod.xlsx")

def cargaExcel(tabla: str, codigo_col: int, descripcion_col: int):
    # Borrar tabla antes de cargar nuevos datos
    cursor.execute(f"DELETE FROM {tabla}")  
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        for fila in hoja.iter_rows(min_row=2, values_only=True):

            codigo = fila[codigo_col]
            descripcion = fila[descripcion_col]
            formulario = hoja.title

            if not codigo or not descripcion or not formulario:
                continue

            try:
                # INSERT dinámico usando f-string para el nombre de la tabla
                cursor.execute(
                    f"INSERT INTO {tabla} (codigo, descripcion, formulario) VALUES (%s, %s, %s)",
                    (codigo, descripcion, formulario)
                )
                conexion.commit()

            except IntegrityError as e:
                if e.errno == 1062:  # codigo duplicado
                    print(f"Duplicado ignorado: {codigo}")
                else:
                    raise
def cargaAutorizacion():
    cursor.execute("TRUNCATE TABLE autorizacion")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            codMEYSS = fila[9]        # Columna A -> COD.MEYSS
            codigo_permiso = fila[10]  # Columna B -> TIPO ACEX
            codigo_via = fila[11]      # Columna C -> VIA ACEX
            formulario = hoja.title

            # Saltar filas incompletas
            if not codMEYSS or not codigo_permiso or not codigo_via:
                print(f"Fila {i} ignorada por valores nulos")
                continue

            # Insertar registro sin filtrar duplicados
            try:
                cursor.execute(
                    "INSERT INTO autorizacion (codMEYSS, codigo_permiso, codigo_via, formulario) VALUES (%s, %s, %s, %s)",
                    (codMEYSS, codigo_permiso, codigo_via, formulario)
                )
            except IntegrityError as e:
                # Esto capturará problemas de FK o cualquier otro error
                print(f"Error insertando {codMEYSS} en fila {i}: {e}")

        conexion.commit()
        print(f"Hoja '{hoja.title}' procesada.")




def menuCargaExcel():
    print("Seleccione una opción:")
    print("1. Cargar Permisos")
    print("2. Cargar Vías")
    print("3. Cargar Autorizaciones")
    print("0. Salir")

    opcion = input("Ingrese el número de la opción: ")

    if opcion == "1":
        cargaExcel("permiso", 10, 3)
    elif opcion == "2":
        cargaExcel("via", 11, 4)
    elif opcion == "3":
        cargaAutorizacion()
    elif opcion == "0":
        print("Saliendo del programa.")
    else:
        print("Opción no válida. Intente de nuevo.")
        menuCargaExcel()
if __name__ == "__main__":
 menuCargaExcel()
