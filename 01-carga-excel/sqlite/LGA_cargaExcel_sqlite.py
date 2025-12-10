import signal, sys
import os
from openpyxl import load_workbook
import sqlite3
from sqlite3 import IntegrityError


def cortarPrograma(sig, frame):
    print("\n\n[!]Saliendo ...\n")
    sys.exit(1)


# Ctlr + C
signal.signal(signal.SIGINT, cortarPrograma)

# Obtener el directorio del script
script_dir = os.path.dirname(os.path.abspath(__file__))
db_dir = os.path.join(script_dir, "db")
docs_dir = os.path.join(script_dir, "docs")

# Variables globales
fichero_permisos = os.path.join(docs_dir, "excel-cod.xlsx")

conn = sqlite3.connect(os.path.join(db_dir, "formulario.sqlite"))
cursor = conn.cursor()
wb = load_workbook(fichero_permisos)

def cargaModelos():
    cursor.execute("DELETE FROM lga_modelos")
    for nombre_hoja in wb.sheetnames:
        id_formulario = nombre_hoja  # el ID viene del nombre de la hoja
        des_modelo = ''               # vacío por ahora

        # Saltar si ID vacío por algún motivo
        if not id_formulario:
            print(f"Hoja ignorada por ID vacío: {nombre_hoja}")
            continue

        try:
            cursor.execute(
                "INSERT INTO LGA_MODELOS (ID, DES_MODELO) VALUES (?, ?)",
                (id_formulario, des_modelo)
            )
        except IntegrityError as e:
            print(f"Error insertando {id_formulario}: {e}")

    conn.commit()
    print("Todos los modelos procesados.")

def cargaPermisos():
    cursor.execute("DELETE FROM lga_permisos")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            id_permiso = fila[10]      
            des_permiso = fila[3]    
            lucrativo = fila[19] 
            residencia = fila[18]
            via_defecto = fila[11]
            meses_validez = fila[7]
            reglamento = fila[12]

            # Saltar filas incompletas
            if not id_permiso or not des_permiso or not lucrativo :
                print(f"Fila {i} ignorada por campos vacíos")
                continue
            def normalizar_valor(valor):
                return 'N' if valor in ("N/A", None) else 'S'
            lucrativo = normalizar_valor(lucrativo)
            residencia = normalizar_valor(residencia)
            try:
                cursor.execute(
                    "INSERT INTO LGA_PERMISOS (ID, DES_PERMISO, LUCRATIVO, RESIDENCIA, VIA_DEFECTO, MESES_VALIDEZ, REGLAMENTO) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (id_permiso, des_permiso, lucrativo,residencia, via_defecto, meses_validez, reglamento)
                )
            except IntegrityError as e:
                print(f"Error insertando {id_permiso} en fila {i}: {e}")

        conn.commit()
        print(f"Hoja '{hoja.title}' procesada.")
def cargaVias():
    #cursor.execute("DELETE FROM lga_via_acceso")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            id_via = fila[11]      
            des_via_acceso = fila[5]

            # Saltar filas incompletas
            if not id_via or not des_via_acceso:
                print(f"Fila {i} ignorada por campos vacíos")
                continue

            try:
                cursor.execute(
                    "INSERT INTO LGA_VIA_ACCESO (ID, DES_VIA_ACCESO) VALUES (?, ?)",
                    (id_via, des_via_acceso)
                )
            except IntegrityError as e:
                print(f"Error insertando {id_via} en fila {i}: {e}")

        conn.commit()
        print(f"Hoja '{hoja.title}' procesada.")
def cargaAutorizaciones():
    cursor.execute("DELETE FROM lga_autorizaciones")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            cod_MEYSS = fila[9]      
            id_permiso = fila[10]
            id_via = fila[11]
            id_modelo = nombre_hoja   

            # Saltar filas incompletas
            if not cod_MEYSS or not id_permiso or not id_via:
                print(f"Fila {i} ignorada por campos vacíos")
                continue

            try:
                # Verificar que las claves foráneas existan antes de insertar
                cursor.execute("SELECT 1 FROM LGA_MODELOS WHERE ID = ?", (id_modelo,))
                if not cursor.fetchone():
                    print(f"Fila {i}: Modelo '{id_modelo}' no existe")
                    continue
                
                cursor.execute("SELECT 1 FROM LGA_PERMISOS WHERE ID = ?", (id_permiso,))
                if not cursor.fetchone():
                    print(f"Fila {i}: Permiso '{id_permiso}' no existe")
                    continue
                
                cursor.execute("SELECT 1 FROM LGA_VIA_ACCESO WHERE ID = ?", (id_via,))
                if not cursor.fetchone():
                    print(f"Fila {i}: Vía '{id_via}' no existe")
                    continue
                
                # Si todas las claves foráneas existen, hacer el INSERT
                cursor.execute(
                    "INSERT INTO LGA_AUTORIZACIONES (COD_MEYSS, ID_PERMISO, ID_VIA, ID_MODELO) VALUES (?, ?, ?, ?)",
                    (cod_MEYSS, id_permiso, id_via, id_modelo)
                )
            except IntegrityError as e:
                print(f"Error insertando {cod_MEYSS} en fila {i}: {e}")

        conn.commit()
        print(f"Hoja '{hoja.title}' procesada.")
def menu():
    print("Seleccione una opción:")
    print("1. Cargar Modelos")
    print("2. Cargar Permisos")
    print("3. Cargar Vias")
    print("4. Cargar Autorizaciones")
    print("0. Salir\n")

def menuExcel():
      while True:
        menu()
        opcion = input("Ingrese el número de la opción: ")
        if opcion == "1": 
            # Cargar Modelos
            cargaModelos()
        elif opcion == "2":
            # Cargar Permisos
            cargaPermisos()
        elif opcion == "3":
            # Cargar Vias
            cargaVias()
        elif opcion == "4":
            # Cargar Autorizaciones
            cargaAutorizaciones()
        elif opcion == "0":
            print("Saliendo del programa.")
            break
        else:
            print("Opción no válida. Intente de nuevo.\n")
            continue

if __name__ == "__main__":
    menuExcel()