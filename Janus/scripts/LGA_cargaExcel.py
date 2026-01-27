import signal, sys, os
import re, unicodedata
import mysql.connector
import openpyxl
from mysql.connector import IntegrityError

def cortarPrograma(sig, frame):
    print("\n\n[!]Saliendo ...\n")
    sys.exit(1)

# Ctlr + C
signal.signal(signal.SIGINT, cortarPrograma)
conexion = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="lga_formularios",
    port=3306,
)

cursor = conexion.cursor()
# Obtener el directorio del script
script_dir = os.path.dirname(os.path.abspath(__file__))
docs_dir = os.path.join(script_dir, "docs")

fichero_permisos = os.path.join(docs_dir, "excel-cod.xlsx")
wb = openpyxl.load_workbook(fichero_permisos)
def cargaModelos():
    cursor.execute("DELETE FROM lga_modelos")
    for nombre_hoja in wb.sheetnames:
        id_formulario = nombre_hoja  # el ID viene del nombre de la hoja
        des_modelo = ''               # vacío por ahora

        try:
            cursor.execute(
                "INSERT INTO LGA_MODELOS (ID, DES_MODELO) VALUES (%s, %s)",
                (id_formulario, des_modelo)
            )
        except IntegrityError as e:
            print(f"Error insertando {id_formulario}: {e}")


    conexion.commit()
    print("Todos los modelos procesados.")

def cargaPermisos():
    cursor.execute("DELETE FROM lga_permisos")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True)):
            id_permiso = fila[10]      
            des_permiso = fila[3]    
            lucrativo = fila[19] 
            residencia = fila[18]
            via_defecto = fila[11]
            reglamento = fila[12]

            def normalizar_valor(valor):
                return 'N' if valor in ("N/A", None) else 'S'
            lucrativo = normalizar_valor(lucrativo)
            residencia = normalizar_valor(residencia)
            #Saltar filas incompletas
            if not id_permiso or not des_permiso or not via_defecto:
                print(f"Fila {i+2} ignorada por campos vacíos")
                continue
            try:
                cursor.execute(
                    "INSERT INTO LGA_PERMISOS (ID, DES_PERMISO, LUCRATIVO, RESIDENCIA, VIA_DEFECTO, REGLAMENTO) VALUES (%s, %s, %s, %s, %s, %s)",
                    (id_permiso, des_permiso, lucrativo, residencia, via_defecto, reglamento)
                )
            except IntegrityError as e:
                print(f"Error insertando {id_permiso} en fila {i+2}: {e}")

        conexion.commit()
        print(f"Hoja '{hoja.title}' procesada.")
def cargaVias():
    cursor.execute("DELETE FROM lga_via_acceso")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True)):
            id_via = fila[11]      
            des_via_acceso = fila[5]
            try:
                #Salta filas incompletas
                if not id_via:
                    print(f"Fila {i+2} ignorada por campos vacíos")
                    continue
                cursor.execute(
                    "INSERT INTO LGA_VIA_ACCESO (ID, DES_VIA_ACCESO) VALUES (%s, %s)",
                    (id_via, des_via_acceso)
                )
            except IntegrityError as e:
                print(f"Error insertando {id_via} en fila {i+2}: {e}")

        conexion.commit()
        print(f"Hoja '{hoja.title}' procesada.")
def cargaAutorizaciones():
    cursor.execute("DELETE FROM lga_autorizaciones")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True)):
            cod_MEYSS = fila[9]      
            id_permiso = fila[10]
            id_via = fila[11]
            num_plazos = fila[7]
            tipo_plazo = fila[7]
            silencio = fila[8]
            tasa_052 = fila[18]
            tasa_062 = fila[19]
            tasa_dos_veces_smi= fila[20]
            autoriza_trabajar= fila[21]
            duracion= fila[22]
            id_modelo = nombre_hoja   

            try:
                #Salta filas incompletas
                if not cod_MEYSS or not id_permiso or not id_via or not id_modelo:
                    print(f"Fila {i+2} ignorada por campos vacíos")
                    continue
                # Validaciones antes de insertar
                cursor.execute("SELECT 1 FROM LGA_MODELOS WHERE ID = %s", (id_modelo,))
                if not cursor.fetchone():
                    print(f"Fila {i+2}: Modelo '{id_modelo}' no existe")
                    continue
                
                cursor.execute("SELECT 1 FROM LGA_PERMISOS WHERE ID = %s", (id_permiso,))
                if not cursor.fetchone():
                    print(f"Fila {i+2}: Permiso '{id_permiso}' no existe")
                    continue
                
                cursor.execute("SELECT 1 FROM LGA_VIA_ACCESO WHERE ID = %s", (id_via,))
                if not cursor.fetchone():
                    print(f"Fila {i+2}: Vía '{id_via}' no existe")
                    continue
                # Extraer solo caracteres numéricos
                num_plazos_str = ''.join(filter(str.isdigit, str(num_plazos)))
                num_plazos = int(num_plazos_str) if num_plazos_str else 0
                # Formatear tipo_plazo
                tipo_plazo = formateo_tipo_plazo(str(tipo_plazo))
                # Si todas las claves foráneas existen, hacer el INSERT
                if silencio not in ('S', 'N'):
                    silencio = 'N'  # Valor por defecto si no es válido:
                patron_dos_veces_smi = r"^.{1}\..{1}\..+"
                if re.match(patron_dos_veces_smi, str(tasa_dos_veces_smi)) is None:
                    tasa_dos_veces_smi = 'N'
                else: 
                    tasa_dos_veces_smi = 'S'
                try:
                    autoriza_trabajar = validar_autoriza_trabajar(str(autoriza_trabajar))
                except ValueError as ve:
                    print(f"Fila {i+2}: {ve}")
                    continue
                duracion = parse_duracion(str(duracion))
                cursor.execute(
                    "INSERT INTO LGA_AUTORIZACIONES (COD_MEYSS, ID_PERMISO, ID_VIA, ID_MODELO, NUM_PLAZO, TIPO_PLAZO, SILENCIO, EPIGRAFE_TASA_052, EPIGRAFE_TASA_062, DOS_VECES_SMI, AUTORIZA_TRABAJAR, DURACION) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (cod_MEYSS, id_permiso, id_via, id_modelo, num_plazos, tipo_plazo, silencio, tasa_052, tasa_062, tasa_dos_veces_smi, autoriza_trabajar, duracion)
                )
            except IntegrityError as e:
                print(f"Error insertando {cod_MEYSS} en fila {i+2}: {e}")

        conexion.commit()
        print(f"Hoja '{hoja.title}' procesada.")

def normalizar_texto(texto: str) -> str:
    texto = texto.strip().upper()
    texto = ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )
    return texto

import re

def parse_duracion(duracion: str):
    """
    Devuelve:
      - 'IN' si es indefinida (MENOS de 1 AÑO, MAX, <, >, HASTA 3 MESES, PLURIANUAL, etc)
      - 'DE' si es desconocida / no especificada
      - int (meses) si es una duración concreta
    """
    if not duracion:
        return "DE"

    txt = normalizar_texto(duracion)

    # Palabras o símbolos que indican rango o indefinido
    patrones_indefinidos = [
        "INDEFINIDO", "LIMITE", "ILIMITADO", "PLURIANUAL",
        "MENOS", "MAX", "MIN", "HASTA", "MAS DE", "SUPERIOR A", "INFERIOR A",
        "<", ">", ">=", "<="
    ]

    for p in patrones_indefinidos:
        if p in txt:
            return "IN"

    # Valor técnico de indefinido
    if "-1" in txt:
        return "IN"

    # Buscar número concreto
    match = re.search(r'(\d+)\s*(DIA|DIAS|ANO|ANOS|MES|MESES)', txt)
    if not match:
        return "DE"

    valor = int(match.group(1))
    unidad = match.group(2)

    if unidad.startswith("ANO"):
        meses = valor * 12
    elif unidad.startswith("DIA"):
        meses = valor // 30  # Aproximación a meses
    else:  # MES / MESES
        meses = valor

    return meses

def validar_autoriza_trabajar(valor: str | None):
    if valor is None or str(valor).strip() == "":
        return None

    texto = normalizar_texto(valor)

    # Casos directos
    if "CUENTA AJENA" in texto:
        return "S"   # Sí autoriza a trabajar

    # Casos ambiguos o combinados
    if texto in ["/", "-", "N/A"]:
        return "A"

    if "C/A" in texto or "C/P" in texto:
        # C/A, C/P, C/A Y C/P → ambiguo
        return "A"

    # Reglas de detección (orden importa)
    reglas = {
        'S': ['SI', 'SÍ', 'AUTORIZA', 'AUTORIZADO', 'TRABAJA'],
        'N': ['NO'],
        'I': ['INSTRUCCION', 'INSTRUCCIONES'],
        'R': ['RESOLUCION'],
        'P': ['PRACTICA', 'PRACTICAS'],
        'L': ['NO AUTORIZA', 'DATOS LABORALES'],
        'A': ['SI/NO', 'S/N']
    }

    for codigo, palabras_clave in reglas.items():
        for palabra in palabras_clave:
            if palabra in texto:
                return codigo

    # Si llega aquí, no se pudo determinar
    raise ValueError(
        f"No se pudo interpretar el valor '{valor}' para 'Autoriza a trabajar'"
    )


def formateo_tipo_plazo(STRING: str) -> str:
    STRING = STRING.strip().casefold()
    if STRING.endswith("meses") or STRING.endswith("mes"):
        return "M"
    elif STRING.endswith("dias") or STRING.endswith("dia"):
        return "D"
    else:
        return "M" # Valor por defecto
def BorrarTablas():
    tablas = ["LGA_MODELOS", "LGA_PERMISOS", "LGA_VIA_ACCESO", "LGA_AUTORIZACIONES"]
    for tabla in tablas:
        cursor.execute(f"DELETE FROM {tabla}")
    conexion.commit()
    print("Todas las tablas han sido borradas.")

def menu():
    print("Seleccione una opción:")
    print("1. Cargar Modelos")
    print("2. Cargar Permisos")
    print("3. Cargar Vias")
    print("4. Cargar Autorizaciones")
    print("5. Borrar Tablas")
    print("0. Salir\n")

def menuExcel():
      while True:
        menu()
        opcion = input("Ingrese el número de la opción: ")
        if opcion == "1": 
            # Cargar Modelos
            cargaModelos()
        elif opcion == "2":
            # Cargar Permisos
            cargaPermisos()
        elif opcion == "3":
            # Cargar Vias
            cargaVias()
        elif opcion == "4":
            # Cargar Autorizaciones
            cargaAutorizaciones()
        elif opcion == "5":
            # Borrar Tablas
            BorrarTablas()
        elif opcion == "0":
            print("Saliendo del programa.")
            break
        else:
            print("Opción no válida. Intente de nuevo.\n")
            continue

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python carga.py [modelos|permisos|vias|autorizaciones|borrar]")
        sys.exit(1)

    opcion = sys.argv[1].lower()

    if opcion == "modelos":
        cargaModelos()
    elif opcion == "permisos":
        cargaPermisos()
    elif opcion == "vias":
        cargaVias()
    elif opcion == "autorizaciones":
        cargaAutorizaciones()
    elif opcion == "borrar":
        BorrarTablas()
    else:
        print("Opción no reconocida:", opcion)
